from openpyxl import load_workbook

g_wb = 0

def pcl_load_workbook(filename:str):
    global g_wb
    g_wb = load_workbook(
        filename=filename,
        read_only=False,
        keep_vba=False,
        keep_links=False)

def pcl_free_workbook():
    global g_wb
    del g_wb

def pcl_get_worksheet(worksheetname):
    global g_wb
    return g_wb[worksheetname]
